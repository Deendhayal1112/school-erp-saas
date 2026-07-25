import uuid
from datetime import date
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.exceptions.exceptions import BadRequestException
from app.models.school import School
from app.modules.student.enums import StudentStatus
from app.modules.student.exceptions import (
    DuplicateAdmissionNumberException,
    DuplicateEmailException,
    InvalidAdmissionDateException,
    InvalidAgeException,
    StudentNotFoundException,
)
from app.modules.student.models import Student
from app.modules.student.repository import StudentRepository
from app.modules.student.schemas import StudentCreate, StudentUpdate
from app.modules.student.validators import validate_student_dob


class StudentService:
    """
    Business service layer applying domain validation rules for Student records.
    """

    def __init__(self, repository: StudentRepository, session: AsyncSession) -> None:
        self.repo = repository
        self.session = session

    async def create_student(self, schema: StudentCreate) -> Student:
        """Applies business validations and persists a new student record."""
        # 1. Validate School tenant exists
        school_stmt = select(School).where(
            School.id == schema.school_id, School.is_deleted == False
        )
        school_res = await self.session.execute(school_stmt)
        if not school_res.scalar_one_or_none():
            raise StudentNotFoundException(
                f"School tenant with id {schema.school_id} not found."
            )

        # 2. Prevent duplicate admission number
        if await self.repo.exists_by_admission_number(
            schema.school_id, schema.admission_number
        ):
            raise DuplicateAdmissionNumberException()

        # 3. Prevent duplicate email
        if schema.email and await self.repo.exists_by_email(
            schema.school_id, schema.email
        ):
            raise DuplicateEmailException()

        # 4. Validate age limits (2 to 30 years old)
        try:
            validate_student_dob(schema.date_of_birth)
        except ValueError as exc:
            raise InvalidAgeException(str(exc))

        # 5. Validate joined and graduation dates
        today = date.today()
        if schema.joined_date > today:
            raise InvalidAdmissionDateException("Joined date cannot be in the future.")
        if schema.graduation_date and schema.graduation_date < schema.joined_date:
            raise InvalidAdmissionDateException(
                "Graduation date must be after joined date."
            )

        # 6. Map to model and persist
        student = Student(
            school_id=schema.school_id,
            admission_number=schema.admission_number,
            roll_number=schema.roll_number,
            emis_number=schema.emis_number,
            first_name=schema.first_name,
            middle_name=schema.middle_name,
            last_name=schema.last_name,
            gender=schema.gender,
            date_of_birth=schema.date_of_birth,
            blood_group=schema.blood_group,
            email=schema.email,
            phone=schema.phone,
            aadhaar_number=schema.aadhaar_number,
            nationality=schema.nationality,
            religion=schema.religion,
            caste=schema.caste,
            community=schema.community,
            mother_tongue=schema.mother_tongue,
            photo_url=schema.photo_url,
            joined_date=schema.joined_date,
            graduation_date=schema.graduation_date,
            remarks=schema.remarks,
            status=StudentStatus.NEW,  # Force initial status to NEW on registration
        )
        result = await self.repo.create(student)
        await (
            self.session.flush()
        )  # Populate DB-generated fields (id, timestamps, defaults)
        return result

    async def update_student(
        self, student_id: uuid.UUID, schema: StudentUpdate
    ) -> Student:
        """Applies mutation validations and updates student record details."""
        student = await self.repo.get_by_id(student_id)
        if not student:
            raise StudentNotFoundException()

        # 1. Prevent duplicate admission number if changed
        if (
            schema.admission_number
            and schema.admission_number != student.admission_number
        ):
            if await self.repo.exists_by_admission_number(
                student.school_id, schema.admission_number
            ):
                raise DuplicateAdmissionNumberException()
            student.admission_number = schema.admission_number

        # 2. Prevent duplicate email if changed
        if schema.email and schema.email != student.email:
            if await self.repo.exists_by_email(student.school_id, schema.email):
                raise DuplicateEmailException()
            student.email = schema.email

        # 3. Validate age limits if DOB changed
        if schema.date_of_birth:
            try:
                validate_student_dob(schema.date_of_birth)
            except ValueError as exc:
                raise InvalidAgeException(str(exc))
            student.date_of_birth = schema.date_of_birth

        # 4. Validate admission overrides
        joined = schema.joined_date or student.joined_date
        grad = (
            schema.graduation_date
            if schema.graduation_date is not None
            else student.graduation_date
        )

        if schema.joined_date and schema.joined_date > date.today():
            raise InvalidAdmissionDateException("Joined date cannot be in the future.")

        if grad and grad < joined:
            raise InvalidAdmissionDateException(
                "Graduation date must be after joined date."
            )

        if schema.joined_date:
            student.joined_date = schema.joined_date
        if schema.graduation_date is not None:
            student.graduation_date = schema.graduation_date

        # 5. Validate status transitions
        if schema.status and schema.status != student.status:
            # Transition back to NEW is prohibited once student has graduated/active/dropped
            if (
                schema.status == StudentStatus.NEW
                and student.status != StudentStatus.NEW
            ):
                raise BadRequestException("Cannot revert student status back to NEW.")
            student.status = schema.status

        # 6. Apply simple field mappings
        if schema.roll_number is not None:
            student.roll_number = schema.roll_number
        if schema.emis_number is not None:
            student.emis_number = schema.emis_number
        if schema.first_name is not None:
            student.first_name = schema.first_name
        if schema.middle_name is not None:
            student.middle_name = schema.middle_name
        if schema.last_name is not None:
            student.last_name = schema.last_name
        if schema.gender is not None:
            student.gender = schema.gender
        if schema.blood_group is not None:
            student.blood_group = schema.blood_group
        if schema.phone is not None:
            student.phone = schema.phone
        if schema.aadhaar_number is not None:
            student.aadhaar_number = schema.aadhaar_number
        if schema.nationality is not None:
            student.nationality = schema.nationality
        if schema.religion is not None:
            student.religion = schema.religion
        if schema.caste is not None:
            student.caste = schema.caste
        if schema.community is not None:
            student.community = schema.community
        if schema.mother_tongue is not None:
            student.mother_tongue = schema.mother_tongue
        if schema.photo_url is not None:
            student.photo_url = schema.photo_url
        if schema.is_active is not None:
            student.is_active = schema.is_active
        if schema.remarks is not None:
            student.remarks = schema.remarks

        result = await self.repo.update(student)
        await self.session.flush()  # Populate updated timestamps
        return result

    async def delete_student(self, student_id: uuid.UUID) -> bool:
        """Deletes a student record (soft-delete)."""
        student = await self.repo.get_by_id(student_id)
        if not student:
            raise StudentNotFoundException()
        return await self.repo.delete(student_id)

    async def restore_student(self, student_id: uuid.UUID) -> bool:
        """Restores a soft-deleted student record."""
        student = await self.repo.get_by_id(student_id, include_deleted=True)
        if not student:
            raise StudentNotFoundException()
        return await self.repo.restore(student_id)

    async def bulk_delete_students(
        self, student_ids: list[uuid.UUID], school_id: uuid.UUID
    ) -> int:
        """Bulk soft-deletes a list of students, isolating by school tenant."""
        count = 0
        for sid in student_ids:
            student = await self.repo.get_by_id(sid)
            if student and student.school_id == school_id:
                if await self.repo.delete(sid):
                    count += 1
        await self.session.flush()
        return count

    async def bulk_restore_students(
        self, student_ids: list[uuid.UUID], school_id: uuid.UUID
    ) -> int:
        """Bulk restores a list of soft-deleted students, isolating by school tenant."""
        count = 0
        for sid in student_ids:
            student = await self.repo.get_by_id(sid, include_deleted=True)
            if student and student.school_id == school_id:
                if await self.repo.restore(sid):
                    count += 1
        await self.session.flush()
        return count

    async def bulk_update_status(
        self, student_ids: list[uuid.UUID], status: StudentStatus, school_id: uuid.UUID
    ) -> int:
        """Bulk updates the status of a list of students, isolating by school tenant."""
        count = 0
        for sid in student_ids:
            student = await self.repo.get_by_id(sid)
            if student and student.school_id == school_id:
                if status == StudentStatus.NEW and student.status != StudentStatus.NEW:
                    # Ignore/skip invalid transitions in bulk update
                    continue
                student.status = status
                self.session.add(student)
                count += 1
        await self.session.flush()
        return count

    async def import_students(
        self, file_content: bytes, filename: str, school_id: uuid.UUID
    ) -> dict[str, Any]:
        """
        Parses and imports students from a CSV or Excel (.xlsx) file with full validation.
        Performs tenant isolation, row-level validation, and duplicate rejection.
        """
        import csv
        import io
        from datetime import date, datetime

        import openpyxl

        rows: list[dict[str, Any]] = []
        ext = filename.split(".")[-1].lower()

        try:
            if ext == "csv":
                text_content = file_content.decode("utf-8-sig")  # utf-8-sig handles BOM
                reader = csv.DictReader(io.StringIO(text_content))
                # Map headers case-insensitively
                fieldnames = reader.fieldnames or []
                header_map = self._normalize_headers(list(fieldnames))
                for row in reader:
                    mapped_row = {
                        header_map[k]: v for k, v in row.items() if k in header_map
                    }
                    rows.append(mapped_row)
            elif ext in ("xlsx", "xls"):
                wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                sheet = wb.active
                if sheet:
                    # Extract headers
                    headers = [str(cell.value or "").strip() for cell in sheet[1]]
                    header_map = self._normalize_headers(headers)
                    # Extract rows
                    for r_idx in range(2, sheet.max_row + 1):
                        row_vals = [
                            sheet.cell(row=r_idx, column=c_idx).value
                            for c_idx in range(1, len(headers) + 1)
                        ]
                        # Check if row is completely empty
                        if not any(v is not None for v in row_vals):
                            continue
                        row_dict = {}
                        for h, val in zip(headers, row_vals, strict=False):
                            if h in header_map:
                                # Convert datetime objects to date for date fields
                                if isinstance(val, (datetime, date)) and header_map[
                                    h
                                ] in (
                                    "date_of_birth",
                                    "joined_date",
                                    "graduation_date",
                                ):
                                    if isinstance(val, datetime):
                                        val = val.date()
                                row_dict[header_map[h]] = val
                        rows.append(row_dict)
            else:
                raise BadRequestException(
                    f"Unsupported file format: {ext}. Only CSV and XLSX are supported."
                )
        except BadRequestException:
            raise
        except Exception as exc:
            raise BadRequestException(f"Failed to parse file: {exc!s}")

        imported_count = 0
        failed_count = 0
        skipped_count = 0
        details: list[dict[str, Any]] = []

        seen_admissions = set()
        seen_emails = set()

        for idx, row in enumerate(rows, start=2):
            errors = []

            # Clean and normalize strings
            for k, v in row.items():
                if isinstance(v, str):
                    row[k] = v.strip() or None

            # Skip empty rows
            if not any(v is not None for v in row.values()):
                skipped_count += 1
                details.append(
                    {
                        "row_number": idx,
                        "admission_number": None,
                        "status": "skipped",
                        "errors": ["Empty row."],
                    }
                )
                continue

            admission_number = row.get("admission_number")
            email = row.get("email")

            # Duplicate checking inside file
            if admission_number:
                if admission_number in seen_admissions:
                    errors.append(
                        f"Duplicate admission number '{admission_number}' in import file."
                    )
                else:
                    seen_admissions.add(admission_number)

            if email:
                if email in seen_emails:
                    errors.append(f"Duplicate email '{email}' in import file.")
                else:
                    seen_emails.add(email)

            # DB level duplicates and format checks
            if admission_number and not errors:
                if await self.repo.exists_by_admission_number(
                    school_id, admission_number
                ):
                    errors.append(
                        f"Admission number '{admission_number}' is already registered in DB."
                    )
            if email and not errors:
                if await self.repo.exists_by_email(school_id, email):
                    errors.append(f"Email '{email}' is already registered in DB.")

            # Map date strings to date objects if needed
            for date_field in ("date_of_birth", "joined_date", "graduation_date"):
                val = row.get(date_field)
                if isinstance(val, str):
                    try:
                        # try YYYY-MM-DD
                        row[date_field] = datetime.strptime(val, "%Y-%m-%d").date()
                    except ValueError:
                        try:
                            # try DD/MM/YYYY
                            row[date_field] = datetime.strptime(val, "%d/%m/%Y").date()
                        except ValueError:
                            try:
                                # try DD-MM-YYYY
                                row[date_field] = datetime.strptime(
                                    val, "%d-%m-%Y"
                                ).date()
                            except ValueError:
                                errors.append(
                                    f"Invalid date format for '{date_field}': '{val}'. Use YYYY-MM-DD."
                                )

            # Validate schemas using Pydantic StudentCreate validator
            if not errors:
                row["school_id"] = school_id
                try:
                    # Enforce fields formatting & validators
                    # If gender is a string, capitalize it to match Enum values
                    if isinstance(row.get("gender"), str):
                        row["gender"] = row["gender"].upper()

                    student_create = StudentCreate.model_validate(row)

                    # Custom date-logical check (joined vs graduation)
                    today = date.today()
                    if student_create.joined_date > today:
                        errors.append("Joined date cannot be in the future.")
                    if (
                        student_create.graduation_date
                        and student_create.graduation_date < student_create.joined_date
                    ):
                        errors.append("Graduation date must be after joined date.")

                except ValueError as err:
                    errors.append(str(err))
                except Exception as err:
                    # Extract custom pydantic validation errors nicely
                    if hasattr(err, "errors"):
                        for e in err.errors():
                            field = " -> ".join(str(loc) for loc in e.get("loc", []))
                            msg = e.get("msg", "Validation error")
                            errors.append(f"{field}: {msg}")
                    else:
                        errors.append(str(err))

            if errors:
                failed_count += 1
                details.append(
                    {
                        "row_number": idx,
                        "admission_number": admission_number,
                        "status": "failed",
                        "errors": errors,
                    }
                )
            else:
                try:
                    # Create student
                    student = Student(
                        school_id=school_id,
                        admission_number=student_create.admission_number,
                        roll_number=student_create.roll_number,
                        emis_number=student_create.emis_number,
                        first_name=student_create.first_name,
                        middle_name=student_create.middle_name,
                        last_name=student_create.last_name,
                        gender=student_create.gender,
                        date_of_birth=student_create.date_of_birth,
                        blood_group=student_create.blood_group,
                        email=student_create.email,
                        phone=student_create.phone,
                        aadhaar_number=student_create.aadhaar_number,
                        nationality=student_create.nationality,
                        religion=student_create.religion,
                        caste=student_create.caste,
                        community=student_create.community,
                        mother_tongue=student_create.mother_tongue,
                        photo_url=student_create.photo_url,
                        joined_date=student_create.joined_date,
                        graduation_date=student_create.graduation_date,
                        remarks=student_create.remarks,
                        status=StudentStatus.NEW,
                    )
                    await self.repo.create(student)
                    imported_count += 1
                    details.append(
                        {
                            "row_number": idx,
                            "admission_number": admission_number,
                            "status": "imported",
                            "errors": [],
                        }
                    )
                except Exception as db_err:
                    failed_count += 1
                    details.append(
                        {
                            "row_number": idx,
                            "admission_number": admission_number,
                            "status": "failed",
                            "errors": [f"Database insert failed: {db_err!s}"],
                        }
                    )

        await self.session.flush()

        return {
            "imported": imported_count,
            "failed": failed_count,
            "skipped": skipped_count,
            "details": details,
        }

    def _normalize_headers(self, headers: list[str]) -> dict[str, str]:
        """Normalizes file headers to match StudentCreate attributes."""
        mapping = {
            "admissionnumber": "admission_number",
            "admission_number": "admission_number",
            "admissionno": "admission_number",
            "admission_no": "admission_number",
            "first_name": "first_name",
            "firstname": "first_name",
            "first name": "first_name",
            "middle_name": "middle_name",
            "middlename": "middle_name",
            "middle name": "middle_name",
            "last_name": "last_name",
            "lastname": "last_name",
            "last name": "last_name",
            "gender": "gender",
            "dateofbirth": "date_of_birth",
            "date_of_birth": "date_of_birth",
            "dob": "date_of_birth",
            "birthdate": "date_of_birth",
            "birth_date": "date_of_birth",
            "email": "email",
            "phone": "phone",
            "phonenumber": "phone",
            "phone_number": "phone",
            "mobile": "phone",
            "contact": "phone",
            "aadhaar": "aadhaar_number",
            "aadhaarnumber": "aadhaar_number",
            "aadhaar_number": "aadhaar_number",
            "aadhar": "aadhaar_number",
            "rollnumber": "roll_number",
            "roll_number": "roll_number",
            "rollno": "roll_number",
            "roll_no": "roll_number",
            "emisnumber": "emis_number",
            "emis_number": "emis_number",
            "emis": "emis_number",
            "joineddate": "joined_date",
            "joined_date": "joined_date",
            "joiningdate": "joined_date",
            "joining_date": "joined_date",
            "admissiondate": "joined_date",
            "bloodgroup": "blood_group",
            "blood_group": "blood_group",
            "nationality": "nationality",
            "religion": "religion",
            "caste": "caste",
            "community": "community",
            "mothertongue": "mother_tongue",
            "mother_tongue": "mother_tongue",
            "photo": "photo_url",
            "photourl": "photo_url",
            "photo_url": "photo_url",
            "remarks": "remarks",
        }
        res = {}
        for h in headers:
            normalized = h.lower().replace(" ", "").replace("_", "").replace("-", "")
            if normalized in mapping:
                res[h] = mapping[normalized]
        return res

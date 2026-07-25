"""
Integration tests for Student advanced capability module.
"""

import io
import uuid
from datetime import date, timedelta

import openpyxl
import pytest
from httpx import ASGITransport, AsyncClient
from sqlalchemy import select

from app.core.password import hash_password
from app.db.session import AsyncSessionLocal
from app.main import app
from app.models.role import Role
from app.models.school import School
from app.models.user import User
from app.modules.student.models import Student

BASE_URL = "/api/v1/students"


@pytest.fixture
async def client() -> AsyncClient:
    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://testserver"
    ) as ac:
        yield ac


@pytest.fixture
async def school_fixtures():
    """Seeds two schools for tenant isolation verification."""
    async with AsyncSessionLocal() as session:
        school1 = School(
            name="Primary Test School",
            code=f"PRM_{uuid.uuid4().hex[:6]}",
            email=f"primary_{uuid.uuid4().hex[:6]}@school.com",
            status="active",
        )
        school2 = School(
            name="Secondary Test School",
            code=f"SEC_{uuid.uuid4().hex[:6]}",
            email=f"secondary_{uuid.uuid4().hex[:6]}@school.com",
            status="active",
        )
        session.add(school1)
        session.add(school2)
        await session.commit()
        await session.refresh(school1)
        await session.refresh(school2)

        yield school1, school2

        # Cleanup
        async with AsyncSessionLocal() as session:
            s1 = await session.get(School, school1.id)
            s2 = await session.get(School, school2.id)
            if s1:
                await session.delete(s1)
            if s2:
                await session.delete(s2)
            await session.commit()


@pytest.fixture
async def auth_headers_prm(client: AsyncClient, school_fixtures) -> dict:
    """Creates a SUPER_ADMIN user for Primary School."""
    school1, _ = school_fixtures
    async with AsyncSessionLocal() as session:
        role_res = await session.execute(select(Role).where(Role.code == "SUPER_ADMIN"))
        role = role_res.scalar_one()

        email = f"prm_admin_{uuid.uuid4().hex[:8]}@test.com"
        username = f"prmadmin_{uuid.uuid4().hex[:8]}"
        pwd = "TestSecret123!"

        user = User(
            first_name="Primary",
            last_name="Admin",
            username=username,
            email=email,
            password_hash=hash_password(pwd),
            school_id=school1.id,
            role_id=role.id,
            status="active",
            email_verified=True,
        )
        session.add(user)
        await session.commit()

    login_resp = await client.post(
        "/api/v1/auth/login",
        json={"email": email, "password": pwd},
    )
    assert login_resp.status_code == 200
    token = login_resp.json()["access_token"]
    yield {"Authorization": f"Bearer {token}"}

    # Cleanup
    async with AsyncSessionLocal() as session:
        user_res = await session.execute(select(User).where(User.email == email))
        u = user_res.scalar_one_or_none()
        if u:
            await session.delete(u)
            await session.commit()


@pytest.fixture
async def auth_headers_sec(client: AsyncClient, school_fixtures) -> dict:
    """Creates a SUPER_ADMIN user for Secondary School."""
    _, school2 = school_fixtures
    async with AsyncSessionLocal() as session:
        role_res = await session.execute(select(Role).where(Role.code == "SUPER_ADMIN"))
        role = role_res.scalar_one()

        email = f"sec_admin_{uuid.uuid4().hex[:8]}@test.com"
        username = f"secadmin_{uuid.uuid4().hex[:8]}"
        pwd = "TestSecret123!"

        user = User(
            first_name="Secondary",
            last_name="Admin",
            username=username,
            email=email,
            password_hash=hash_password(pwd),
            school_id=school2.id,
            role_id=role.id,
            status="active",
            email_verified=True,
        )
        session.add(user)
        await session.commit()

    login_resp = await client.post(
        "/api/v1/auth/login",
        json={"email": email, "password": pwd},
    )
    assert login_resp.status_code == 200
    token = login_resp.json()["access_token"]
    yield {"Authorization": f"Bearer {token}"}

    # Cleanup
    async with AsyncSessionLocal() as session:
        user_res = await session.execute(select(User).where(User.email == email))
        u = user_res.scalar_one_or_none()
        if u:
            await session.delete(u)
            await session.commit()


# Helper to create payload
def _payload(school_id: uuid.UUID, **overrides) -> dict:
    p = {
        "school_id": str(school_id),
        "admission_number": f"ADM_{uuid.uuid4().hex[:8]}",
        "first_name": "Integration",
        "last_name": "Test",
        "gender": "MALE",
        "date_of_birth": "2015-06-15",
        "joined_date": str(date.today() - timedelta(days=30)),
        "status": "NEW",
    }
    p.update(overrides)
    return p


@pytest.mark.asyncio
async def test_search_and_filters_and_sorting(
    client: AsyncClient, auth_headers_prm: dict, school_fixtures
):
    """Verifies advanced search, combinable dynamic filtering, and multi-column sorting."""
    school1, _ = school_fixtures

    # Seed 3 students with distinct details
    s1_payload = _payload(
        school1.id,
        first_name="Alice",
        middle_name="Grace",
        last_name="Smith",
        gender="FEMALE",
        blood_group="A+",
        joined_date="2024-01-10",
        admission_number="ADM_ALICE",
        phone="+1234567890",
    )
    s2_payload = _payload(
        school1.id,
        first_name="Bob",
        last_name="Smith",
        gender="MALE",
        blood_group="O-",
        joined_date="2024-02-15",
        admission_number="ADM_BOB",
        phone="+1987654321",
    )
    s3_payload = _payload(
        school1.id,
        first_name="Charlie",
        last_name="Brown",
        gender="MALE",
        blood_group="A+",
        joined_date="2024-03-20",
        admission_number="ADM_CHARLIE",
        phone="+1555555555",
    )

    for p in (s1_payload, s2_payload, s3_payload):
        resp = await client.post(BASE_URL + "/", json=p, headers=auth_headers_prm)
        assert resp.status_code == 201

    try:
        # 1. Advanced Search on middle name
        resp = await client.get(BASE_URL + "/?search=Grace", headers=auth_headers_prm)
        assert resp.status_code == 200, f"Search failed: {resp.text}"
        results = resp.json()["results"]
        assert len(results) == 1, f"Expected 1, got {len(results)}. Results: {results}"
        assert results[0]["first_name"] == "Alice"

        # Search on phone number
        resp = await client.get(
            BASE_URL + "/?search=1987654321", headers=auth_headers_prm
        )
        assert resp.status_code == 200
        results = resp.json()["results"]
        assert len(results) == 1
        assert results[0]["first_name"] == "Bob"

        # 2. Dynamic Filtering: Combined filters
        resp = await client.get(
            BASE_URL + "/?gender=MALE&blood_group=A%2B", headers=auth_headers_prm
        )
        assert resp.status_code == 200
        results = resp.json()["results"]
        assert len(results) == 1
        assert results[0]["first_name"] == "Charlie"

        # Joined date range filters
        resp = await client.get(
            BASE_URL + "/?joined_date_from=2024-01-01&joined_date_to=2024-02-28",
            headers=auth_headers_prm,
        )
        assert resp.status_code == 200
        results = resp.json()["results"]
        first_names = {r["first_name"] for r in results}
        assert "Alice" in first_names
        assert "Bob" in first_names
        assert "Charlie" not in first_names

        # 3. Multi-column Sorting (first name ascending, last name descending)
        resp = await client.get(
            BASE_URL + "/?sort=first_name,-last_name", headers=auth_headers_prm
        )
        assert resp.status_code == 200
        results = resp.json()["results"]
        # Filter only Alice, Bob, Charlie (to skip potential residual seed data)
        sorted_names = [
            r["first_name"]
            for r in results
            if r["first_name"] in ("Alice", "Bob", "Charlie")
        ]
        assert sorted_names == ["Alice", "Bob", "Charlie"]

    finally:
        # Cleanup
        async with AsyncSessionLocal() as session:
            res = await session.execute(
                select(Student).where(Student.school_id == school1.id)
            )
            for s in res.scalars():
                await session.delete(s)
            await session.commit()


@pytest.mark.asyncio
async def test_bulk_operations(
    client: AsyncClient, auth_headers_prm: dict, school_fixtures
):
    """Verifies POST /bulk-delete, bulk-restore, and bulk-status."""
    school1, _ = school_fixtures

    # Seed two students
    s1 = _payload(school1.id, admission_number="BULK_1")
    s2 = _payload(school1.id, admission_number="BULK_2")

    r1 = await client.post(BASE_URL + "/", json=s1, headers=auth_headers_prm)
    r2 = await client.post(BASE_URL + "/", json=s2, headers=auth_headers_prm)
    assert r1.status_code == 201
    assert r2.status_code == 201

    id1 = r1.json()["data"]["id"]
    id2 = r2.json()["data"]["id"]

    try:
        # 1. Bulk Update Status to ACTIVE
        resp = await client.post(
            BASE_URL + "/bulk-status",
            json={"student_ids": [id1, id2], "status": "ACTIVE"},
            headers=auth_headers_prm,
        )
        assert resp.status_code == 200
        assert resp.json()["data"]["count"] == 2

        # Verify status in database
        g1 = await client.get(f"{BASE_URL}/{id1}", headers=auth_headers_prm)
        g2 = await client.get(f"{BASE_URL}/{id2}", headers=auth_headers_prm)
        assert g1.json()["data"]["status"] == "ACTIVE"
        assert g2.json()["data"]["status"] == "ACTIVE"

        # Reverting to NEW bulk status transition test
        resp = await client.post(
            BASE_URL + "/bulk-status",
            json={"student_ids": [id1], "status": "NEW"},
            headers=auth_headers_prm,
        )
        assert resp.status_code == 200
        # Revert back to NEW is prohibited by service and skipped/ignored inside bulk
        assert resp.json()["data"]["count"] == 0

        # 2. Bulk Delete (Soft-Delete)
        resp = await client.post(
            BASE_URL + "/bulk-delete",
            json={"student_ids": [id1, id2]},
            headers=auth_headers_prm,
        )
        assert resp.status_code == 200
        assert resp.json()["data"]["count"] == 2

        # Verify soft-deleted students are not retrievable via GET
        g1 = await client.get(f"{BASE_URL}/{id1}", headers=auth_headers_prm)
        assert g1.status_code == 404

        # 3. Bulk Restore
        resp = await client.post(
            BASE_URL + "/bulk-restore",
            json={"student_ids": [id1, id2]},
            headers=auth_headers_prm,
        )
        assert resp.status_code == 200
        assert resp.json()["data"]["count"] == 2

        # Verify retrieved again
        g1 = await client.get(f"{BASE_URL}/{id1}", headers=auth_headers_prm)
        assert g1.status_code == 200

    finally:
        async with AsyncSessionLocal() as session:
            res = await session.execute(
                select(Student).where(Student.school_id == school1.id)
            )
            for s in res.scalars():
                await session.delete(s)
            await session.commit()


@pytest.mark.asyncio
async def test_csv_import_success_and_failures(
    client: AsyncClient, auth_headers_prm: dict, school_fixtures
):
    """Verifies POST /import parsing and detailed validation errors on CSV."""
    school1, _ = school_fixtures

    csv_data = (
        "admission_number,first_name,last_name,gender,date_of_birth,joined_date,email,phone,aadhaar_number\n"
        "IMP_CSV_1,John,Doe,Male,2015-05-10,2023-06-01,john.doe@csv.com,+1234567890,123456789012\n"  # Success
        "IMP_CSV_2,Jane,Doe,Female,2015-05-10,2023-06-01,jane.doe@csv.com,invalid_phone,123456789012\n"  # Invalid phone
        "IMP_CSV_1,Duplicate,Admission,Male,2015-05-10,2023-06-01,dup@csv.com,+1234567890,123456789012\n"  # Duplicate admission
        "IMP_CSV_3,Invalid,Dob,Male,2099-05-10,2023-06-01,dob@csv.com,+1234567890,123456789012\n"  # DOB future
    )

    try:
        files = {"file": ("students.csv", csv_data, "text/csv")}
        resp = await client.post(
            BASE_URL + "/import", files=files, headers=auth_headers_prm
        )
        assert resp.status_code == 201

        data = resp.json()["data"]
        assert data["imported"] == 1
        assert data["failed"] == 3
        assert data["skipped"] == 0

        # Verify failures detail
        details = {d["row_number"]: d for d in data["details"]}
        assert details[3]["status"] == "failed"
        assert any("Phone" in err or "phone" in err for err in details[3]["errors"])

        assert details[4]["status"] == "failed"
        assert any(
            "Duplicate" in err or "duplicate" in err for err in details[4]["errors"]
        )

        assert details[5]["status"] == "failed"
        assert any("future" in err for err in details[5]["errors"])

    finally:
        async with AsyncSessionLocal() as session:
            res = await session.execute(
                select(Student).where(Student.school_id == school1.id)
            )
            for s in res.scalars():
                await session.delete(s)
            await session.commit()


@pytest.mark.asyncio
async def test_excel_import_success_and_failures(
    client: AsyncClient, auth_headers_prm: dict, school_fixtures
):
    """Verifies POST /import parsing and validation errors on Excel (.xlsx)."""
    school1, _ = school_fixtures

    # Build XLSX workbook in memory using openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = [
        "admission_number",
        "first_name",
        "last_name",
        "gender",
        "date_of_birth",
        "joined_date",
        "email",
        "phone",
        "aadhaar_number",
    ]
    ws.append(headers)

    # Rows
    ws.append(
        [
            "IMP_XLSX_1",
            "Alex",
            "Lee",
            "Male",
            "2015-05-10",
            "2023-06-01",
            "alex@xlsx.com",
            "+1234567890",
            "123456789012",
        ]
    )  # Success
    ws.append(
        [
            "IMP_XLSX_2",
            "Betty",
            "Lee",
            "Female",
            "2015-05-10",
            "2023-06-01",
            "betty@xlsx.com",
            "invalid_phone",
            "123456789012",
        ]
    )  # Invalid phone
    ws.append(
        [
            "IMP_XLSX_1",
            "Duplicate",
            "Alex",
            "Male",
            "2015-05-10",
            "2023-06-01",
            "dup@xlsx.com",
            "+1234567890",
            "123456789012",
        ]
    )  # Duplicate admission

    stream = io.BytesIO()
    wb.save(stream)
    excel_bytes = stream.getvalue()

    try:
        files = {
            "file": (
                "students.xlsx",
                excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        resp = await client.post(
            BASE_URL + "/import", files=files, headers=auth_headers_prm
        )
        assert resp.status_code == 201

        data = resp.json()["data"]
        assert data["imported"] == 1
        assert data["failed"] == 2
        assert data["skipped"] == 0

    finally:
        async with AsyncSessionLocal() as session:
            res = await session.execute(
                select(Student).where(Student.school_id == school1.id)
            )
            for s in res.scalars():
                await session.delete(s)
            await session.commit()


@pytest.mark.asyncio
async def test_bulk_and_filtered_exports(
    client: AsyncClient, auth_headers_prm: dict, school_fixtures
):
    """Verifies POST /bulk-export and GET /export."""
    school1, _ = school_fixtures

    # Create 2 students
    s1 = _payload(school1.id, first_name="Export1")
    s2 = _payload(school1.id, first_name="Export2")

    r1 = await client.post(BASE_URL + "/", json=s1, headers=auth_headers_prm)
    r2 = await client.post(BASE_URL + "/", json=s2, headers=auth_headers_prm)
    assert r1.status_code == 201
    id1 = r1.json()["data"]["id"]
    id2 = r2.json()["data"]["id"]

    try:
        # 1. Bulk export POST (Excel)
        resp = await client.post(
            BASE_URL + "/bulk-export?format=excel",
            json={"student_ids": [id1, id2]},
            headers=auth_headers_prm,
        )
        assert resp.status_code == 200
        assert (
            resp.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # Parse Excel output from bytes
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
        sheet = wb.active
        assert sheet.cell(row=2, column=4).value in ("Export1", "Export2")

        # 2. Filtered list export GET (CSV)
        resp = await client.get(
            BASE_URL + "/export?format=csv&search=Export1",
            headers=auth_headers_prm,
        )
        assert resp.status_code == 200, f"Export failed: {resp.text}"
        assert resp.headers["content-type"] == "text/csv; charset=utf-8"
        assert "Export1" in resp.text
        assert "Export2" not in resp.text

        # 3. PDF placeholder export GET (PDF)
        resp = await client.get(
            BASE_URL + "/export?format=pdf",
            headers=auth_headers_prm,
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"] == "application/pdf"
        assert resp.content.startswith(b"%PDF-1.4")

    finally:
        async with AsyncSessionLocal() as session:
            res = await session.execute(
                select(Student).where(Student.school_id == school1.id)
            )
            for s in res.scalars():
                await session.delete(s)
            await session.commit()


@pytest.mark.asyncio
async def test_tenant_isolation(
    client: AsyncClient, auth_headers_prm: dict, auth_headers_sec: dict, school_fixtures
):
    """Verifies that user context school_id isolation prevents viewing/modifying cross-tenant data."""
    school1, _ = school_fixtures

    # Seed student in primary school A
    payload_a = _payload(school1.id, first_name="SchoolAStudent")
    resp_a = await client.post(BASE_URL + "/", json=payload_a, headers=auth_headers_prm)
    assert resp_a.status_code == 201
    id_a = resp_a.json()["data"]["id"]

    try:
        # User B queries list -> Should NOT see Student A
        list_b = await client.get(BASE_URL + "/", headers=auth_headers_sec)
        assert list_b.status_code == 200
        results = list_b.json()["results"]
        assert not any(r["id"] == id_a for r in results)

        # User B queries single -> Should receive 404
        get_b = await client.get(f"{BASE_URL}/{id_a}", headers=auth_headers_sec)
        assert get_b.status_code == 404

        # User B updates -> Should receive 404
        up_b = await client.put(
            f"{BASE_URL}/{id_a}", json={"first_name": "Hack"}, headers=auth_headers_sec
        )
        assert up_b.status_code == 404

        # User B deletes -> Should receive 404
        del_b = await client.delete(f"{BASE_URL}/{id_a}", headers=auth_headers_sec)
        assert del_b.status_code == 404

        # User B bulk status updates -> Should modify 0 records
        bulk_status = await client.post(
            BASE_URL + "/bulk-status",
            json={"student_ids": [id_a], "status": "ACTIVE"},
            headers=auth_headers_sec,
        )
        assert bulk_status.status_code == 200
        assert bulk_status.json()["data"]["count"] == 0

        # User B bulk deletes -> Should delete 0 records
        bulk_del = await client.post(
            BASE_URL + "/bulk-delete",
            json={"student_ids": [id_a]},
            headers=auth_headers_sec,
        )
        assert bulk_del.status_code == 200
        assert bulk_del.json()["data"]["count"] == 0

    finally:
        async with AsyncSessionLocal() as session:
            res = await session.execute(
                select(Student).where(Student.school_id == school1.id)
            )
            for s in res.scalars():
                await session.delete(s)
            await session.commit()
